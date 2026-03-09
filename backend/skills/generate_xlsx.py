#!/usr/bin/env python3
"""Generate XLSX files from a JSON specification.

Takes a JSON file describing sheets, data, formulas, formatting, and metadata,
and produces a fully-formatted .xlsx file with real Excel formulas.

Usage:
  python scripts/generate_xlsx.py INPUT.json OUTPUT.xlsx
  python scripts/generate_xlsx.py --validate INPUT.json
  python scripts/generate_xlsx.py --verify OUTPUT.xlsx

Prerequisites:
  .claude/pyproject.toml — must declare openpyxl dependency

Run via:
  uv run --directory .claude python .claude/skills/spreadsheet-generation/scripts/generate_xlsx.py spec.json output.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# JSON schema definition for input validation
# ---------------------------------------------------------------------------

REQUIRED_TOP_KEYS = {"sheets"}
VALID_TOP_KEYS = {"sheets", "named_ranges", "metadata"}
VALID_SHEET_KEYS = {
    "name", "data", "formulas", "column_widths", "row_heights",
    "merge_cells", "freeze_panes", "comments", "data_validation",
    "conditional_formatting", "number_formats", "header_rows",
    "auto_filter", "cell_formats",
}
VALID_FORMAT_KEYS = {
    "bold", "italic", "font_size", "font_color", "bg_color",
    "number_format", "alignment", "border",
}


def validate_spec(spec: dict) -> list[str]:
    """Validate a workbook spec dict. Returns list of error strings (empty = valid)."""
    errors = []

    missing = REQUIRED_TOP_KEYS - set(spec.keys())
    if missing:
        errors.append(f"Missing required top-level keys: {missing}")

    unknown = set(spec.keys()) - VALID_TOP_KEYS
    if unknown:
        errors.append(f"Unknown top-level keys: {unknown}")

    sheets = spec.get("sheets", [])
    if not isinstance(sheets, list) or len(sheets) == 0:
        errors.append("'sheets' must be a non-empty list")
        return errors

    sheet_names = set()
    for i, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            errors.append(f"Sheet {i}: must be an object")
            continue
        name = sheet.get("name")
        if not name:
            errors.append(f"Sheet {i}: missing 'name'")
        elif name in sheet_names:
            errors.append(f"Sheet {i}: duplicate name '{name}'")
        else:
            sheet_names.add(name)

        unknown_sk = set(sheet.keys()) - VALID_SHEET_KEYS
        if unknown_sk:
            errors.append(f"Sheet '{name}': unknown keys {unknown_sk}")

        # data must be list of lists
        data = sheet.get("data", [])
        if not isinstance(data, list):
            errors.append(f"Sheet '{name}': 'data' must be a list of rows")

        # formulas must be dict of cell_ref -> formula_string
        formulas = sheet.get("formulas", {})
        if not isinstance(formulas, dict):
            errors.append(f"Sheet '{name}': 'formulas' must be an object {{cell: formula}}")

    return errors


# ---------------------------------------------------------------------------
# Workbook generation
# ---------------------------------------------------------------------------

def build_workbook(spec: dict):
    """Build an openpyxl Workbook from a validated spec dict."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    sheets_spec = spec["sheets"]

    for sheet_spec in sheets_spec:
        ws = wb.create_sheet(title=sheet_spec["name"])

        # --- Write data rows ---
        for row_idx, row_data in enumerate(sheet_spec.get("data", []), start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # --- Write formulas ---
        for cell_ref, formula in sheet_spec.get("formulas", {}).items():
            ws[cell_ref] = formula

        # --- Column widths ---
        for col_spec in sheet_spec.get("column_widths", []):
            col_letter = col_spec["column"]
            ws.column_dimensions[col_letter].width = col_spec["width"]

        # --- Row heights ---
        for rh in sheet_spec.get("row_heights", []):
            ws.row_dimensions[rh["row"]].height = rh["height"]

        # --- Header row formatting ---
        for hr in sheet_spec.get("header_rows", []):
            row_num = hr["row"]
            fmt = hr.get("format", {})
            font_kwargs = {}
            if fmt.get("bold"):
                font_kwargs["bold"] = True
            if fmt.get("italic"):
                font_kwargs["italic"] = True
            if fmt.get("font_size"):
                font_kwargs["size"] = fmt["font_size"]
            if fmt.get("font_color"):
                font_kwargs["color"] = fmt["font_color"]
            font = Font(**font_kwargs) if font_kwargs else None

            fill = None
            if fmt.get("bg_color"):
                fill = PatternFill(start_color=fmt["bg_color"],
                                   end_color=fmt["bg_color"],
                                   fill_type="solid")

            alignment = None
            if fmt.get("alignment"):
                alignment = Alignment(**fmt["alignment"])

            for cell in ws[row_num]:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment

        # --- Cell-level formatting (IB color coding, per-cell font/bg) ---
        for cf in sheet_spec.get("cell_formats", []):
            cf_font_kwargs = {}
            if cf.get("bold"):
                cf_font_kwargs["bold"] = True
            if cf.get("italic"):
                cf_font_kwargs["italic"] = True
            if cf.get("font_size"):
                cf_font_kwargs["size"] = cf["font_size"]
            if cf.get("font_color"):
                cf_font_kwargs["color"] = cf["font_color"]
            cf_font = Font(**cf_font_kwargs) if cf_font_kwargs else None

            cf_fill = None
            if cf.get("bg_color"):
                cf_fill = PatternFill(start_color=cf["bg_color"],
                                      end_color=cf["bg_color"],
                                      fill_type="solid")

            target_cells = list(cf.get("cells", []))
            for range_str in cf.get("ranges", []):
                for row in ws[range_str]:
                    if hasattr(row, '__iter__'):
                        for cell in row:
                            target_cells.append(cell.coordinate)
                    else:
                        target_cells.append(row.coordinate)

            for cell_ref in target_cells:
                cell = ws[cell_ref]
                if cf_font:
                    existing = cell.font
                    merged = Font(
                        bold=cf_font.bold if cf_font.bold else existing.bold,
                        italic=cf_font.italic if cf_font.italic else existing.italic,
                        size=cf_font.size if cf_font.size else existing.size,
                        color=cf_font.color if cf_font.color else existing.color,
                    )
                    cell.font = merged
                if cf_fill:
                    cell.fill = cf_fill

        # --- Number formats ---
        for nf in sheet_spec.get("number_formats", []):
            cells = nf.get("cells", [])
            ranges = nf.get("ranges", [])
            fmt_str = nf["format"]

            for cell_ref in cells:
                ws[cell_ref].number_format = fmt_str

            for range_str in ranges:
                for row in ws[range_str]:
                    if hasattr(row, '__iter__'):
                        for cell in row:
                            cell.number_format = fmt_str
                    else:
                        row.number_format = fmt_str

        # --- Merge cells ---
        for merge_range in sheet_spec.get("merge_cells", []):
            ws.merge_cells(merge_range)

        # --- Freeze panes ---
        freeze = sheet_spec.get("freeze_panes")
        if freeze:
            ws.freeze_panes = freeze

        # --- Comments ---
        for comment_spec in sheet_spec.get("comments", []):
            cell_ref = comment_spec["cell"]
            text = comment_spec["text"]
            author = comment_spec.get("author", "EduPitch")
            ws[cell_ref].comment = Comment(text, author)

        # --- Data validation ---
        for dv_spec in sheet_spec.get("data_validation", []):
            dv = DataValidation(
                type=dv_spec.get("type", "list"),
                formula1=dv_spec.get("formula1"),
                formula2=dv_spec.get("formula2"),
                allow_blank=dv_spec.get("allow_blank", True),
            )
            if dv_spec.get("prompt_title"):
                dv.prompt = dv_spec.get("prompt_message", "")
                dv.promptTitle = dv_spec["prompt_title"]
                dv.showInputMessage = True
            if dv_spec.get("error_title"):
                dv.error = dv_spec.get("error_message", "")
                dv.errorTitle = dv_spec["error_title"]
                dv.showErrorMessage = True
            dv.sqref = dv_spec["sqref"]
            ws.add_data_validation(dv)

        # --- Conditional formatting ---
        for cf_spec in sheet_spec.get("conditional_formatting", []):
            range_str = cf_spec["range"]
            rule_type = cf_spec["rule_type"]

            if rule_type == "cell_is":
                rule = CellIsRule(
                    operator=cf_spec["operator"],
                    formula=cf_spec.get("formula", []),
                    font=Font(color=cf_spec.get("font_color")) if cf_spec.get("font_color") else None,
                    fill=PatternFill(
                        start_color=cf_spec.get("bg_color", "FFFFFF"),
                        end_color=cf_spec.get("bg_color", "FFFFFF"),
                        fill_type="solid",
                    ) if cf_spec.get("bg_color") else None,
                )
                ws.conditional_formatting.add(range_str, rule)
            elif rule_type == "color_scale":
                rule = ColorScaleRule(
                    start_type=cf_spec.get("start_type", "min"),
                    start_color=cf_spec.get("start_color", "F8696B"),
                    end_type=cf_spec.get("end_type", "max"),
                    end_color=cf_spec.get("end_color", "63BE7B"),
                )
                ws.conditional_formatting.add(range_str, rule)
            elif rule_type == "data_bar":
                rule = DataBarRule(
                    start_type=cf_spec.get("start_type", "min"),
                    end_type=cf_spec.get("end_type", "max"),
                    color=cf_spec.get("color", "638EC6"),
                )
                ws.conditional_formatting.add(range_str, rule)

        # --- Auto filter ---
        auto_filter = sheet_spec.get("auto_filter")
        if auto_filter:
            ws.auto_filter.ref = auto_filter

    # --- Named ranges ---
    for nr in spec.get("named_ranges", []):
        from openpyxl.workbook.defined_name import DefinedName
        defn = DefinedName(nr["name"], attr_text=nr["value"])
        wb.defined_names.add(defn)

    # --- Metadata ---
    meta = spec.get("metadata", {})
    if meta.get("title"):
        wb.properties.title = meta["title"]
    if meta.get("creator"):
        wb.properties.creator = meta["creator"]
    if meta.get("description"):
        wb.properties.description = meta["description"]

    return wb


def verify_xlsx(filepath: str) -> dict:
    """Open an xlsx file and report basic stats for verification."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=False)
    result = {
        "file": filepath,
        "sheets": [],
        "named_ranges": list(wb.defined_names.keys()),
    }

    for ws in wb:
        sheet_info = {
            "name": ws.title,
            "dimensions": ws.dimensions,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "formula_cells": [],
            "comment_cells": [],
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        }
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_info["formula_cells"].append(
                        f"{cell.coordinate}: {cell.value}"
                    )
                if cell.comment:
                    sheet_info["comment_cells"].append(
                        f"{cell.coordinate}: {cell.comment.text[:80]}"
                    )
        result["sheets"].append(sheet_info)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate XLSX from JSON spec, or validate/verify files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s spec.json output.xlsx          Generate XLSX from spec
  %(prog)s --validate spec.json           Validate spec without generating
  %(prog)s --verify output.xlsx           Verify an existing XLSX file
        """,
    )
    parser.add_argument("input", nargs="?", help="Input JSON spec file")
    parser.add_argument("output", nargs="?", help="Output XLSX file path")
    parser.add_argument("--validate", metavar="FILE", help="Validate a JSON spec (no generation)")
    parser.add_argument("--verify", metavar="FILE", help="Verify an existing XLSX file")

    args = parser.parse_args()

    # Mode: validate
    if args.validate:
        spec = json.loads(Path(args.validate).read_text())
        errors = validate_spec(spec)
        if errors:
            print("VALIDATION FAILED:")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)
        else:
            print("VALIDATION OK")
            print(f"  Sheets: {len(spec['sheets'])}")
            for s in spec["sheets"]:
                rows = len(s.get("data", []))
                formulas = len(s.get("formulas", {}))
                print(f"    - {s['name']}: {rows} data rows, {formulas} formulas")
            sys.exit(0)

    # Mode: verify
    if args.verify:
        result = verify_xlsx(args.verify)
        print(json.dumps(result, indent=2))
        sys.exit(0)

    # Mode: generate
    if not args.input or not args.output:
        parser.error("Provide INPUT.json and OUTPUT.xlsx, or use --validate/--verify")

    spec = json.loads(Path(args.input).read_text())

    # Validate first
    errors = validate_spec(spec)
    if errors:
        print("VALIDATION FAILED — cannot generate:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(1)

    wb = build_workbook(spec)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    # Auto-verify
    result = verify_xlsx(str(output_path))
    print(f"Generated: {output_path}")
    print(f"  Sheets: {len(result['sheets'])}")
    for s in result["sheets"]:
        print(f"    - {s['name']}: {s['rows']}x{s['columns']}, "
              f"{len(s['formula_cells'])} formulas, "
              f"{len(s['comment_cells'])} comments")
    if result["named_ranges"]:
        print(f"  Named ranges: {len(result['named_ranges'])}")


if __name__ == "__main__":
    main()
