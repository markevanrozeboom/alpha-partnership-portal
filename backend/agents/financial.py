"""Financial Modelling Agent — investment-bank-quality financial model with interactive assumptions.

Post-workshop (March 16, 2026): Unified two-prong model for all sovereign nations.
No tiers. No PPP scaling. Fixed upfront fees. One model.

Phase 1: Generates configurable assumptions (sliders) from the strategy/country data.
Phase 2: Builds the full P&L, capital deployment, unit economics, returns analysis.
Recalculate: Deterministic recalculation when user adjusts sliders.
"""

from __future__ import annotations

import logging
import math
import os

from models.schemas import (
    FinancialAssumption, FinancialAssumptions,
    FinancialModel, YearProjection, UnitEconomics, CapitalDeployment,
    ReturnsAnalysis, SensitivityScenario,
    Strategy, CountryProfile,
    FlagshipMarketData, MetroFlagshipInput,
    FlagshipOptimizationResult, FlagshipMetroResult,
)
from config import OUTPUT_DIR
from config.rules_loader import (
    get_flagship_tuition_range,
    get_national_per_student_budget,
    get_min_student_year_commit,
    get_fixed_development_costs,
    get_fee_floors,
    get_esa_data,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Flagship Revenue-Maximizing Grid Search (financial_rules_v1.md)
# ---------------------------------------------------------------------------

def _interpolate_eligible_children(
    metro: MetroFlagshipInput,
    min_agi: float,
) -> int:
    """Estimate K-12 children in families with income ≥ min_agi.

    Uses a Pareto (power-law) interpolation between the two known
    data points: children at $200K and children at $500K thresholds.

    Pareto model: N(x) = N_low × (x_low / x)^α
    where α = ln(N_low / N_high) / ln(x_high / x_low)
    """
    n_200k = metro.children_in_families_income_above_200k
    n_500k = metro.children_in_families_income_above_500k

    # Guard: if no data, return 0
    if n_200k <= 0:
        return 0
    # If asking for ≤ $200K threshold, return the $200K count
    if min_agi <= 200_000:
        return n_200k
    # If asking for ≥ $500K threshold, return the $500K count
    if min_agi >= 500_000:
        return max(0, n_500k)

    # Pareto interpolation
    if n_500k <= 0 or n_500k >= n_200k:
        # Can't compute alpha; linear fallback
        frac = (min_agi - 200_000) / (500_000 - 200_000)
        return max(0, int(n_200k + (n_500k - n_200k) * frac))

    alpha = math.log(n_200k / n_500k) / math.log(500_000 / 200_000)
    estimated = n_200k * (200_000 / min_agi) ** alpha
    return max(0, int(estimated))


def optimize_flagships(
    market_data: FlagshipMarketData | None,
    tuition_min: int = 40_000,
    tuition_max: int = 100_000,
    tuition_step: int = 5_000,
    capacity_min: int = 250,
    capacity_max: int = 1_000,
    capacity_step: int = 50,
    penetration_rate: float = 0.20,
) -> FlagshipOptimizationResult:
    """Revenue-maximizing grid search per financial_rules_v1.md.

    For each (tuition, capacity) pair, calculate total revenue across
    all qualifying metros and pick the combination that maximizes it.

    Rules:
    - Max 3 flagships in the capital city, 1 in each other top metro
    - Top metros limited to 3 largest, each must support ≥ 250 students
      at the minimum $40K tuition
    - Tuition must exceed most expensive non-boarding school
    - 20% penetration rate of eligible families
    - Eligible = K-12 children in families with AGI ≥ 5× tuition
    """
    result = FlagshipOptimizationResult()

    if not market_data or not market_data.metros:
        logger.warning(
            "No flagship market data — returning empty result. "
            "market_data=%s, metros=%s",
            "present" if market_data else "None",
            len(market_data.metros) if market_data else 0,
        )
        return result

    logger.info(
        "optimize_flagships: %d metros provided, "
        "country top school $%s (%s)",
        len(market_data.metros),
        f"{market_data.country_most_expensive_nonboarding_tuition:,.0f}",
        market_data.country_most_expensive_nonboarding_school,
    )
    for i, m in enumerate(market_data.metros):
        logger.info(
            "  Input metro %d: %s (capital=%s) — "
            ">$200K=%s children, >$500K=%s children",
            i + 1, m.metro_name, m.is_capital,
            f"{m.children_in_families_income_above_200k:,}",
            f"{m.children_in_families_income_above_500k:,}",
        )

    # Store school comparison data
    result.most_expensive_school_tuition = (
        market_data.country_most_expensive_nonboarding_tuition
    )
    result.most_expensive_school_name = (
        market_data.country_most_expensive_nonboarding_school
    )

    # --- Step 1: Filter qualifying metros ---
    # Each metro must have ≥ 250 eligible children at minimum tuition
    min_agi_at_floor = 5 * tuition_min  # $200K
    qualifying_metros: list[MetroFlagshipInput] = []
    for metro in market_data.metros[:3]:  # Limited to top 3
        eligible = _interpolate_eligible_children(metro, min_agi_at_floor)
        demand = int(eligible * penetration_rate)
        logger.info(
            "  Qualifying check: %s — eligible_children=%s, "
            "demand_at_20%%=%d, min_capacity=%d → %s",
            metro.metro_name, f"{eligible:,}", demand, capacity_min,
            "QUALIFIES" if demand >= capacity_min else "DOES NOT QUALIFY",
        )
        if demand >= capacity_min:
            qualifying_metros.append(metro)

    logger.info(
        "  %d of %d metros qualify for flagships",
        len(qualifying_metros), len(market_data.metros[:3]),
    )

    if not qualifying_metros:
        result.scholarship_needed = True
        # Calculate the gap
        if market_data.metros:
            best_metro = market_data.metros[0]
            eligible = _interpolate_eligible_children(
                best_metro, min_agi_at_floor,
            )
            demand = int(eligible * penetration_rate)
            gap = capacity_min - demand
            result.scholarship_note = (
                f"No metro supports a {capacity_min}-student, "
                f"${tuition_min:,}/year flagship. "
                f"{best_metro.metro_name} has ~{demand:,} eligible "
                f"students at ${tuition_min:,} tuition. "
                f"The country/state would need to fund ~{gap:,} "
                f"scholarships to reach the minimum requirement."
            )
        return result

    # --- Step 2: Determine minimum tuition floor ---
    # Must exceed the most expensive non-boarding school
    school_tuition_floor = (
        market_data.country_most_expensive_nonboarding_tuition
    )
    # Round up to next $5K increment
    if school_tuition_floor > 0:
        effective_min = (
            math.ceil(school_tuition_floor / tuition_step) * tuition_step
        )
        # Must strictly exceed, so bump if equal
        if effective_min <= school_tuition_floor:
            effective_min += tuition_step
    else:
        effective_min = tuition_min
    effective_min = max(effective_min, tuition_min)

    # --- Step 3: Grid search over (tuition, capacity) ---
    best_revenue = 0
    best_tuition = effective_min
    best_capacity = capacity_min
    best_metro_configs: list[dict] = []

    for tuition in range(effective_min, tuition_max + 1, tuition_step):
        min_agi = 5 * tuition

        for capacity in range(capacity_min, capacity_max + 1, capacity_step):
            total_revenue = 0
            metro_configs: list[dict] = []

            for metro in qualifying_metros:
                max_schools = 3 if metro.is_capital else 1
                eligible = _interpolate_eligible_children(metro, min_agi)
                demand = int(eligible * penetration_rate)

                if demand < capacity:
                    # Not enough demand for even 1 school at this config
                    continue

                schools = min(max_schools, demand // capacity)
                if schools < 1:
                    continue

                rev = schools * capacity * tuition
                total_revenue += rev
                metro_configs.append({
                    "metro": metro.metro_name,
                    "is_capital": metro.is_capital,
                    "schools": schools,
                    "capacity": capacity,
                    "tuition": tuition,
                    "revenue": rev,
                    "eligible": eligible,
                    "demand": demand,
                })

            if total_revenue > best_revenue:
                best_revenue = total_revenue
                best_tuition = tuition
                best_capacity = capacity
                best_metro_configs = metro_configs

    # --- Step 4: Build result ---
    result.optimal_tuition = best_tuition
    result.optimal_capacity = best_capacity
    result.total_annual_revenue = best_revenue
    result.tuition_exceeds_most_expensive = (
        best_tuition > school_tuition_floor if school_tuition_floor > 0
        else True
    )

    total_schools = 0
    total_students = 0
    for cfg in best_metro_configs:
        metro_students = cfg["schools"] * cfg["capacity"]
        total_schools += cfg["schools"]
        total_students += metro_students
        result.metros.append(FlagshipMetroResult(
            metro_name=cfg["metro"],
            is_capital=cfg["is_capital"],
            schools=cfg["schools"],
            capacity_per_school=cfg["capacity"],
            tuition=cfg["tuition"],
            annual_revenue=cfg["revenue"],
            eligible_children=cfg["eligible"],
            demand_at_penetration=cfg["demand"],
        ))

    result.total_schools = total_schools
    result.total_students = total_students

    logger.info(
        "Flagship optimization: tuition=$%s, capacity=%d, "
        "schools=%d, students=%d, revenue=$%sM",
        f"{best_tuition:,.0f}", best_capacity,
        total_schools, total_students,
        f"{best_revenue / 1_000_000:,.0f}",
    )

    return result

# ---------------------------------------------------------------------------
# Phase 1: Generate Assumptions
# ---------------------------------------------------------------------------


def generate_assumptions(
    target: str,
    country_profile: CountryProfile,
    strategy: Strategy,
) -> FinancialAssumptions:
    """Generate a set of configurable financial assumptions based on the strategy.

    Sovereign nations use the two-prong model (Flagship + National).
    US states use the existing ESA/school-type model.
    No tiers. No PPP scaling.
    """

    is_us_state = country_profile.target.type.value == "us_state"

    if is_us_state:
        return _generate_us_state_assumptions(target, country_profile, strategy)

    return _generate_sovereign_assumptions(target, country_profile, strategy)


def _generate_sovereign_assumptions(
    target: str,
    country_profile: CountryProfile,
    strategy: Strategy,
) -> FinancialAssumptions:
    """Generate two-prong model assumptions for sovereign nations.

    Prong 1 (Flagship): 100% Alpha-owned. Tuition $40K-$100K, top metros, 25% margin.
    Prong 2 (National): 100% Country-owned, Alpha operates. FIXED $25K/student, 100K min.

    Upfront fees per financial_rules_v1.md:
      4 × $250M fixed development = $1B
      Timeback prepaid = students × $5,000/student
      Operating Fee prepaid = students × $2,500/student
    """
    # --- Load fixed parameters from config ---
    tuition_min, tuition_max = get_flagship_tuition_range()
    national_budget = get_national_per_student_budget()
    min_student_year = get_min_student_year_commit()
    dev_costs = get_fixed_development_costs()
    fee_floors = get_fee_floors()

    # --- Prong 1: Flagship (revenue-maximizing grid search) ---
    flagship_opt = optimize_flagships(
        market_data=country_profile.flagship_market_data,
        tuition_min=tuition_min,
        tuition_max=tuition_max,
    )

    if flagship_opt.total_schools > 0:
        default_flagship_tuition = int(flagship_opt.optimal_tuition)
        default_flagship_schools = flagship_opt.total_schools
        default_flagship_students_per_school = flagship_opt.optimal_capacity
    else:
        # Fallback when no metro-level research data is available yet:
        # Use GDP-per-capita proxy for tuition and sensible defaults for
        # schools / capacity so the term sheet is not empty.
        gdp_cap = country_profile.economy.gdp_per_capita or 30_000
        gdp_ratio = min(1.0, max(0.0, (gdp_cap - 10_000) / 80_000))
        default_flagship_tuition = round(
            (tuition_min + (tuition_max - tuition_min) * gdp_ratio) / 5_000
        ) * 5_000
        default_flagship_tuition = max(
            tuition_min, min(tuition_max, default_flagship_tuition),
        )
        # Reasonable defaults: 3 schools (capital max) × 500 students
        default_flagship_schools = 3
        default_flagship_students_per_school = 500

    # --- Prong 2: National ---
    # 100K student-year minimum commitment
    default_national_students = min_student_year

    # --- Prepaid fees: per-student minimums (NOT revenue %) ---
    # Per financial_rules_v1.md:
    #   Timeback prepaid = students × $5,000 per student
    #   Operating Fee prepaid = students × $2,500 per student
    timeback_floor_per_student = fee_floors.get("timeback_fee_floor_per_student", 5_000)
    mgmt_floor_per_student = fee_floors.get("management_fee_floor_per_student", 2_500)
    prepaid_timeback = round(default_national_students * timeback_floor_per_student / 1_000_000)
    prepaid_mgmt = round(default_national_students * mgmt_floor_per_student / 1_000_000)

    # Fixed development total: $1B (4 × $250M)
    fixed_dev_total_m = round(dev_costs["total"] / 1_000_000)

    # Build description strings from optimization
    _opt_desc_tuition = (
        f"Optimized via grid search. $5K increments. "
        f"Exceeds most expensive non-boarding school "
        f"(${flagship_opt.most_expensive_school_tuition:,.0f})."
        if flagship_opt.most_expensive_school_tuition > 0
        else "Set by AGI of top 20% families. $5K increments."
    )
    _opt_desc_schools = (
        "From grid search: "
        + ", ".join(
            f"{m.metro_name} ({m.schools})"
            for m in flagship_opt.metros
        )
        if flagship_opt.metros
        else "Max 3 in capital + 1 per other top metro"
    )

    assumptions = [
        # --- Prong 1: Flagship ---
        FinancialAssumption(
            key="flagship_tuition", label="Flagship Tuition (per student/year)",
            value=default_flagship_tuition,
            min_val=tuition_min, max_val=tuition_max, step=5_000,
            unit="$", category="prong_1_flagship",
            description=_opt_desc_tuition,
        ),
        FinancialAssumption(
            key="flagship_schools", label="Flagship School Count",
            value=default_flagship_schools,
            min_val=0, max_val=10, step=1,
            unit="schools", category="prong_1_flagship",
            description=_opt_desc_schools,
        ),
        FinancialAssumption(
            key="flagship_students_per_school",
            label="Flagship Students per School",
            value=default_flagship_students_per_school,
            min_val=250, max_val=1_000, step=50,
            unit="students", category="prong_1_flagship",
            description="250-1,000 students, 50-student increments per rules",
        ),
        FinancialAssumption(
            key="flagship_fill_rate", label="Flagship Fill Rate (%)",
            value=50, min_val=30, max_val=100, step=5,
            unit="%", category="prong_1_flagship",
            description="50% backstop guaranteed by country/state for 5 years",
        ),

        # --- Prong 2: National (Counterparty-Owned, Alpha-Operated) ---
        FinancialAssumption(
            key="national_per_student_budget",
            label="National Per-Student Budget",
            value=national_budget,
            min_val=national_budget, max_val=national_budget, step=1_000,
            unit="$", category="prong_2_national",
            description="FIXED $25,000 per student — non-negotiable",
            locked=True,
        ),
        FinancialAssumption(
            key="national_students", label="National Students (Year 5 target)",
            value=default_national_students,
            min_val=min_student_year, max_val=500_000, step=10_000,
            unit="students", category="prong_2_national",
            description=f"Minimum {min_student_year:,} student-year commit",
        ),

        # --- Scale / Ramp ---
        FinancialAssumption(
            key="national_ramp_years", label="Years to Full National Deployment",
            value=3, min_val=2, max_val=5, step=1,
            unit="years", category="scale",
            description="Phased rollout timeline for national prong",
        ),
        FinancialAssumption(
            key="avg_students_per_school", label="Avg Students per School (National)",
            value=800, min_val=200, max_val=2000, step=50,
            unit="students", category="scale",
        ),

        # --- Costs ---
        FinancialAssumption(
            key="cogs_pct", label="COGS (% of Revenue)",
            value=55, min_val=40, max_val=75, step=1,
            unit="%", category="costs",
            description="Direct costs including teacher salaries, facilities, technology",
        ),
        FinancialAssumption(
            key="opex_pct", label="OpEx (% of Revenue)",
            value=20, min_val=10, max_val=35, step=1,
            unit="%", category="costs",
            description="G&A, marketing, central management overhead",
        ),
        FinancialAssumption(
            key="capex_per_school", label="CapEx per New School ($)",
            value=5_000_000, min_val=1_000_000, max_val=20_000_000, step=500_000,
            unit="$", category="costs",
        ),

        # --- Alpha Fee Structure (locked) ---
        FinancialAssumption(
            key="management_fee_pct",
            label="Operating Fee (% of funding/tuition)",
            value=10, min_val=10, max_val=10, step=1,
            unit="%", category="fees",
            description="10% of funding/tuition, min $2,500/student — non-negotiable",
            locked=True,
        ),
        FinancialAssumption(
            key="timeback_license_pct",
            label="Timeback License (% of funding/tuition)",
            value=20, min_val=20, max_val=20, step=1,
            unit="%", category="fees",
            description="20% of funding/tuition, min $5,000/student — non-negotiable",
            locked=True,
        ),

        # --- Upfront Fees: Fixed Development Costs (locked) ---
        FinancialAssumption(
            key="upfront_alphacore_license", label="Alpha Core License ($M)",
            value=dev_costs["alphacore_license"] / 1_000_000,
            min_val=250, max_val=250, step=1,
            unit="$M", category="fees",
            description="FIXED $250M — Alpha Core License, paid upfront",
            locked=True,
        ),
        FinancialAssumption(
            key="upfront_incept_edllm", label="Incept EdLLM ($M)",
            value=dev_costs["incept_edllm"] / 1_000_000,
            min_val=250, max_val=250, step=1,
            unit="$M", category="fees",
            description="FIXED $250M — Country/State specific Incept EdLLM, paid upfront",
            locked=True,
        ),
        FinancialAssumption(
            key="upfront_app_content_rd",
            label="Country-Specific EdTech Apps ($M)",
            value=dev_costs["edtech_app_content_rd"] / 1_000_000,
            min_val=250, max_val=250, step=1,
            unit="$M", category="fees",
            description="FIXED $250M — country-specific EdTech Apps R&D, paid upfront",
            locked=True,
        ),
        FinancialAssumption(
            key="upfront_lifeskills_rd", label="Programs and Life Skills R&D ($M)",
            value=dev_costs["lifeskills_rd"] / 1_000_000,
            min_val=250, max_val=250, step=1,
            unit="$M", category="fees",
            description="FIXED $250M — country-specific Programs and Life Skills, paid upfront",
            locked=True,
        ),

        # --- Prepaid Fees (scale by student count × per-student minimum) ---
        FinancialAssumption(
            key="upfront_timeback_fee", label="Prepaid Timeback Fee ($M)",
            value=prepaid_timeback,
            min_val=1, max_val=5_000, step=5,
            unit="$M", category="fees",
            description=f"Students × ${timeback_floor_per_student:,}/student = ${prepaid_timeback:,}M",
        ),
        FinancialAssumption(
            key="upfront_mgmt_fee", label="Prepaid Operating Fee ($M)",
            value=prepaid_mgmt,
            min_val=1, max_val=2_500, step=5,
            unit="$M", category="fees",
            description=f"Students × ${mgmt_floor_per_student:,}/student = ${prepaid_mgmt:,}M",
        ),
        FinancialAssumption(
            key="parent_education_annual", label="Parent Education / Launch / Guides ($M/yr)",
            value=50, min_val=25, max_val=100, step=5,
            unit="$M", category="fees",
            description="$50M per year ongoing — parent education, marketing, guide training",
        ),
        FinancialAssumption(
            key="upfront_ip_fee", label="Total Upfront Ask ($M)",
            value=fixed_dev_total_m + prepaid_timeback + prepaid_mgmt,
            min_val=1_000, max_val=5_000, step=5,
            unit="$M", category="fees",
            description=(
                f"${fixed_dev_total_m:,}M fixed dev"
                f" + ${prepaid_timeback:,}M Timeback"
                f" + ${prepaid_mgmt:,}M Operating Fee"
            ),
        ),

        # --- Returns ---
        FinancialAssumption(
            key="exit_ebitda_multiple", label="Exit EBITDA Multiple",
            value=15, min_val=8, max_val=25, step=1,
            unit="x", category="returns",
        ),
        FinancialAssumption(
            key="discount_rate", label="Discount Rate (%)",
            value=12, min_val=8, max_val=20, step=1,
            unit="%", category="returns",
        ),
        FinancialAssumption(
            key="tax_rate", label="Tax Rate (%)",
            value=20, min_val=0, max_val=35, step=1,
            unit="%", category="returns",
        ),
    ]

    return FinancialAssumptions(
        assumptions=assumptions,
        flagship_optimization=flagship_opt,
    )


def _generate_us_state_assumptions(
    target: str,
    country_profile: CountryProfile,
    strategy: Strategy,
) -> FinancialAssumptions:
    """Generate assumptions for US states — ESA/voucher model, unchanged from prior logic."""

    # Load ESA-specific data from config if available
    esa_cfg = get_esa_data(target)
    if esa_cfg:
        esa_amount = esa_cfg.get("esa_amount", 8_000)
        base_per_student = max(esa_amount, 18_000)
        premium_tuition = max(25_000, esa_amount * 3)
        mid_tuition = max(15_000, esa_amount * 2)
        logger.info("Using ESA data for %s: $%s/student", target, esa_amount)
    else:
        base_per_student = 18_000
        premium_tuition = 25_000
        mid_tuition = 15_000

    # Student Target (5yr): must be >= 10% of school-age population
    school_age_pop = (
        country_profile.demographics.population_0_18
        or country_profile.education.k12_enrolled
        or 500_000
    )
    min_y5_from_market = int(school_age_pop * 0.10)
    target_students_y5 = strategy.target_student_count_year5 or max(5_000, min_y5_from_market)
    target_students_y5 = max(target_students_y5, min_y5_from_market)

    assumptions = [
        # --- Pricing ---
        FinancialAssumption(
            key="premium_tuition", label="Premium Tier Tuition (per student/year)",
            value=premium_tuition, min_val=15_000, max_val=50_000, step=500,
            unit="$", category="pricing",
            description="Annual tuition for flagship/premium schools",
        ),
        FinancialAssumption(
            key="mid_tuition", label="Mid-Tier Tuition (per student/year)",
            value=mid_tuition, min_val=10_000, max_val=35_000, step=500,
            unit="$", category="pricing",
            description="Annual tuition for mid-market schools",
        ),
        FinancialAssumption(
            key="per_student_budget", label="Per-Student Delivery Budget",
            value=base_per_student, min_val=5_000, max_val=30_000, step=500,
            unit="$", category="pricing",
            description="Per-student operating budget",
        ),

        # --- Scale ---
        FinancialAssumption(
            key="students_year1", label="Year 1 Students",
            value=min(5_000, target_students_y5 // 10),
            min_val=500, max_val=20_000, step=500,
            unit="students", category="scale",
            description="Students in flagship school(s) at launch",
        ),
        FinancialAssumption(
            key="students_year2", label="Year 2 Students",
            value=min(15_000, target_students_y5 // 5),
            min_val=2_000, max_val=50_000, step=1_000,
            unit="students", category="scale",
        ),
        FinancialAssumption(
            key="students_year3", label="Year 3 Students",
            value=min(50_000, target_students_y5 * 2 // 5),
            min_val=5_000, max_val=100_000, step=5_000,
            unit="students", category="scale",
        ),
        FinancialAssumption(
            key="students_year4", label="Year 4 Students",
            value=min(100_000, target_students_y5 * 3 // 5),
            min_val=10_000, max_val=200_000, step=5_000,
            unit="students", category="scale",
        ),
        FinancialAssumption(
            key="students_year5", label="Year 5 Students",
            value=target_students_y5,
            min_val=max(20_000, min_y5_from_market),
            max_val=max(500_000, min_y5_from_market * 5),
            step=10_000,
            unit="students", category="scale",
            description=f"Min {min_y5_from_market:,} (10% of {school_age_pop:,} school-age pop)",
        ),
        FinancialAssumption(
            key="avg_students_per_school", label="Avg Students per School",
            value=800, min_val=200, max_val=2000, step=50,
            unit="students", category="scale",
        ),

        # --- Costs ---
        FinancialAssumption(
            key="cogs_pct", label="COGS (% of Revenue)",
            value=55, min_val=40, max_val=75, step=1,
            unit="%", category="costs",
            description="Direct costs including teacher salaries, facilities, technology",
        ),
        FinancialAssumption(
            key="opex_pct", label="OpEx (% of Revenue)",
            value=20, min_val=10, max_val=35, step=1,
            unit="%", category="costs",
            description="G&A, marketing, central management overhead",
        ),
        FinancialAssumption(
            key="capex_per_school", label="CapEx per New School ($)",
            value=5_000_000, min_val=1_000_000, max_val=20_000_000, step=500_000,
            unit="$", category="costs",
        ),

        # --- Alpha Fee Structure ---
        FinancialAssumption(
            key="management_fee_pct", label="Management Fee (% of School Revenue)",
            value=10, min_val=5, max_val=20, step=1,
            unit="%", category="fees",
            description="Alpha's management fee",
            locked=True,
        ),
        FinancialAssumption(
            key="timeback_license_pct",
            label="Timeback License (% of Per-Student Budget)",
            value=20, min_val=10, max_val=30, step=1,
            unit="%", category="fees",
            description="Alpha's technology/IP licensing fee",
            locked=True,
        ),

        # --- Upfront Payment Breakdown ---
        FinancialAssumption(
            key="upfront_alphacore_license", label="AlphaCore License ($M)",
            value=250, min_val=50, max_val=500, step=10,
            unit="$M", category="fees",
            description="AlphaCore curriculum OS & LMS license — paid to Alpha Holdings",
        ),
        FinancialAssumption(
            key="upfront_mgmt_fee", label="Upfront Management Fee ($M)",
            value=max(1, round(target_students_y5 * base_per_student * 0.10 / 1_000_000)),
            min_val=1, max_val=500, step=5,
            unit="$M", category="fees",
            description="Target Students Y5 × Per-Student Budget × 10%",
        ),
        FinancialAssumption(
            key="upfront_timeback_fee", label="Upfront Timeback Fee ($M)",
            value=max(1, round(target_students_y5 * base_per_student * 0.20 / 1_000_000)),
            min_val=1, max_val=1000, step=5,
            unit="$M", category="fees",
            description="Target Students Y5 × Per-Student Budget × 20%",
        ),
        FinancialAssumption(
            key="upfront_app_content_rd",
            label="Customized State-Specific App Content R&D ($M)",
            value=250, min_val=50, max_val=500, step=10,
            unit="$M", category="fees",
            description="State-specific EdTech app content R&D",
        ),
        FinancialAssumption(
            key="upfront_lifeskills_rd",
            label="Customized State-Specific LifeSkills R&D ($M)",
            value=250, min_val=50, max_val=500, step=10,
            unit="$M", category="fees",
            description="State-specific life skills curriculum R&D",
        ),
        FinancialAssumption(
            key="upfront_ip_fee", label="Total Upfront Ask ($M)",
            value=max(25, 250 + 250 + 250
                      + round(target_students_y5 * base_per_student * 0.10 / 1_000_000)
                      + round(target_students_y5 * base_per_student * 0.20 / 1_000_000)),
            min_val=25, max_val=2000, step=5,
            unit="$M", category="fees",
            description="AlphaCore + App R&D + LifeSkills R&D + Mgmt Fee + Timeback",
        ),
        FinancialAssumption(
            key="guide_school_fee", label="Guide School Training Fee (per teacher, $)",
            value=15_000, min_val=5_000, max_val=30_000, step=1_000,
            unit="$", category="fees",
        ),

        # --- Returns ---
        FinancialAssumption(
            key="exit_ebitda_multiple", label="Exit EBITDA Multiple",
            value=15, min_val=8, max_val=25, step=1,
            unit="x", category="returns",
        ),
        FinancialAssumption(
            key="discount_rate", label="Discount Rate (%)",
            value=12, min_val=8, max_val=20, step=1,
            unit="%", category="returns",
        ),
        FinancialAssumption(
            key="tax_rate", label="Tax Rate (%)",
            value=20, min_val=0, max_val=35, step=1,
            unit="%", category="returns",
        ),
    ]

    return FinancialAssumptions(assumptions=assumptions)


# ---------------------------------------------------------------------------
# Phase 2 / Recalculate: Build model from assumptions
# ---------------------------------------------------------------------------

def build_model(
    assumptions: FinancialAssumptions,
    target: str,
    strategy: Strategy,
) -> FinancialModel:
    """Deterministic financial model calculation from assumptions."""
    a = {item.key: item.value for item in assumptions.assumptions}

    # Detect sovereign two-prong model vs US state model
    if "flagship_tuition" in a:
        return _build_sovereign_model(a, assumptions.flagship_optimization)

    return _build_us_state_model(a)


def _build_sovereign_model(
    a: dict,
    flagship_opt: FlagshipOptimizationResult | None = None,
) -> FinancialModel:
    """Build the two-prong sovereign nation model.

    Prong 1: Flagship schools × students × tuition × fill rate
    Prong 2: National students × $25K
    Alpha revenue: 10% mgmt + 20% timeback on combined revenue
    """
    # --- Extract assumptions ---
    flagship_tuition = a.get("flagship_tuition", 60_000)
    flagship_schools = int(a.get("flagship_schools", 3))
    flagship_per_school = int(a.get("flagship_students_per_school", 800))
    flagship_fill_pct = a.get("flagship_fill_rate", 50) / 100

    national_budget = a.get("national_per_student_budget", 25_000)
    national_students_y5 = int(a.get("national_students", 100_000))
    national_ramp_years = int(a.get("national_ramp_years", 3))

    avg_per_school = int(a.get("avg_students_per_school", 800))
    cogs_pct = a.get("cogs_pct", 55) / 100
    opex_pct = a.get("opex_pct", 20) / 100
    mgmt_fee_pct = a.get("management_fee_pct", 10) / 100
    timeback_pct = a.get("timeback_license_pct", 20) / 100

    capex_per_school = a.get("capex_per_school", 5_000_000)
    exit_multiple = a.get("exit_ebitda_multiple", 15)
    a.get("discount_rate", 12) / 100
    tax_rate = a.get("tax_rate", 20) / 100

    # --- Upfront fees ---
    upfront_alphacore = a.get("upfront_alphacore_license", 250) * 1_000_000
    upfront_incept_edllm = a.get("upfront_incept_edllm", 250) * 1_000_000
    upfront_app_rd = a.get("upfront_app_content_rd", 250) * 1_000_000
    upfront_lifeskills = a.get("upfront_lifeskills_rd", 250) * 1_000_000
    upfront_mgmt = a.get("upfront_mgmt_fee", 250) * 1_000_000      # Operating Fee prepaid
    upfront_timeback = a.get("upfront_timeback_fee", 500) * 1_000_000  # Timeback prepaid
    upfront_ip = a.get("upfront_ip_fee", 1_750) * 1_000_000

    # --- Prong 1: Flagship capacity (constant across years) ---
    flagship_capacity = flagship_schools * flagship_per_school
    flagship_students = int(flagship_capacity * flagship_fill_pct)

    # --- Prong 2: National ramp ---
    def national_students_for_year(yr: int) -> int:
        if yr <= national_ramp_years:
            frac = yr / national_ramp_years
            return max(
                int(national_students_y5 * 0.1),
                int(national_students_y5 * frac),
            )
        return national_students_y5

    # --- P&L Projections ---
    projections: list[YearProjection] = []
    cumulative_cash = 0.0
    prev_schools = 0
    total_mgmt_rev = 0.0
    total_timeback_rev = 0.0

    for yr in range(1, 6):
        nat_students = national_students_for_year(yr)
        total_students = flagship_students + nat_students

        # Revenue
        prong1_rev = flagship_students * flagship_tuition
        prong2_rev = nat_students * national_budget
        revenue = prong1_rev + prong2_rev

        # Schools: flagship are separate, national schools scale
        national_schools = max(1, math.ceil(nat_students / avg_per_school))
        schools = flagship_schools + national_schools
        new_schools = max(0, schools - prev_schools)

        cogs = revenue * cogs_pct
        gross_margin = revenue - cogs
        opex = revenue * opex_pct
        ebitda = gross_margin - opex
        net_income = ebitda * (1 - tax_rate)

        capex = new_schools * capex_per_school
        ip_cost = upfront_ip if yr == 1 else 0
        fcf = net_income - capex - ip_cost
        cumulative_cash += fcf

        # Alpha revenue from this year
        mgmt_rev = revenue * mgmt_fee_pct
        timeback_rev = revenue * timeback_pct
        total_mgmt_rev += mgmt_rev
        total_timeback_rev += timeback_rev

        projections.append(YearProjection(
            year=yr, students=total_students, schools=schools,
            revenue=round(revenue), cogs=round(cogs),
            gross_margin=round(gross_margin), opex=round(opex),
            ebitda=round(ebitda), net_income=round(net_income),
            free_cash_flow=round(fcf), cumulative_cash=round(cumulative_cash),
        ))
        prev_schools = schools

    # --- Unit Economics ---
    unit_econ = [
        UnitEconomics(
            school_type="Flagship (Prong 1)",
            per_student_revenue=flagship_tuition,
            per_student_cost=round(flagship_tuition * cogs_pct),
            contribution_margin=round(flagship_tuition * (1 - cogs_pct)),
            margin_pct=round((1 - cogs_pct) * 100, 1),
        ),
        UnitEconomics(
            school_type="National (Prong 2)",
            per_student_revenue=national_budget,
            per_student_cost=round(national_budget * cogs_pct),
            contribution_margin=round(national_budget * (1 - cogs_pct)),
            margin_pct=round((1 - cogs_pct) * 100, 1),
        ),
    ]

    # --- Capital Deployment ---
    cap_deploy: list[CapitalDeployment] = []
    for yr_idx in range(5):
        p = projections[yr_idx]
        prev_s = projections[yr_idx - 1].schools if yr_idx > 0 else 0
        new_schools_yr = max(0, p.schools - prev_s)

        ip_dev = upfront_ip if yr_idx == 0 else 0
        mgmt = p.revenue * mgmt_fee_pct
        launch = new_schools_yr * capex_per_school

        cap_deploy.append(CapitalDeployment(
            year=yr_idx + 1, ip_development=round(ip_dev),
            management_fees=round(mgmt), launch_capital=round(launch),
            real_estate=round(launch * 0.6),
            total=round(ip_dev + mgmt + launch),
        ))

    # --- Returns Analysis ---
    y5_ebitda = projections[4].ebitda if projections else 0
    ev_exit = y5_ebitda * exit_multiple
    total_invested = sum(cd.total for cd in cap_deploy)
    moic = ev_exit / total_invested if total_invested > 0 else 0

    cash_flows = [-cap_deploy[i].total for i in range(5)]
    cash_flows[4] += ev_exit
    irr = _approx_irr(cash_flows)

    payback = None
    for p in projections:
        if p.cumulative_cash > 0:
            payback = float(p.year)
            break

    returns = ReturnsAnalysis(
        irr=round(irr * 100, 1) if irr else None,
        moic=round(moic, 2),
        enterprise_value_at_exit=round(ev_exit),
        payback_period_years=payback,
        ebitda_multiple=exit_multiple,
    )

    # --- Sensitivity ---
    projections[4].revenue if projections else 0
    sensitivity = [
        SensitivityScenario(
            variable="National Students (Y5)",
            base_case=national_students_y5,
            downside=national_students_y5 * 0.7,
            upside=national_students_y5 * 1.3,
        ),
        SensitivityScenario(
            variable="Flagship Tuition",
            base_case=flagship_tuition,
            downside=flagship_tuition * 0.85,
            upside=flagship_tuition * 1.15,
        ),
        SensitivityScenario(
            variable="COGS %",
            base_case=cogs_pct * 100,
            downside=(cogs_pct + 0.05) * 100,
            upside=(cogs_pct - 0.05) * 100,
        ),
        SensitivityScenario(
            variable="Exit Multiple",
            base_case=exit_multiple,
            downside=exit_multiple * 0.7,
            upside=exit_multiple * 1.3,
        ),
    ]

    return FinancialModel(
        pnl_projection=projections,
        unit_economics=unit_econ,
        capital_deployment=cap_deploy,
        returns_analysis=returns,
        sensitivity=sensitivity,
        management_fee_pct=mgmt_fee_pct,
        timeback_license_pct=timeback_pct,
        upfront_ip_fee=upfront_ip,
        upfront_alphacore_license=upfront_alphacore,
        upfront_incept_edllm=upfront_incept_edllm,
        upfront_app_content_rd=upfront_app_rd,
        upfront_lifeskills_rd=upfront_lifeskills,
        upfront_mgmt_fee=upfront_mgmt,
        upfront_timeback_fee=upfront_timeback,
        total_management_fee_revenue=round(total_mgmt_rev),
        total_timeback_license_revenue=round(total_timeback_rev),
        flagship_tuition=flagship_tuition,
        flagship_students=flagship_students,
        flagship_revenue=round(flagship_students * flagship_tuition),
        flagship_optimization=flagship_opt,
        national_per_student_budget=national_budget,
        national_students=national_students_y5,
        national_revenue=round(national_students_y5 * national_budget),
    )


def _build_us_state_model(a: dict) -> FinancialModel:
    """Build the US state financial model — preserves existing logic."""
    premium_tuition = a.get("premium_tuition", 25_000)
    mid_tuition = a.get("mid_tuition", 15_000)
    per_student_budget = a.get("per_student_budget", 20_000)
    cogs_pct = a.get("cogs_pct", 55) / 100
    opex_pct = a.get("opex_pct", 20) / 100
    mgmt_fee_pct = a.get("management_fee_pct", 10) / 100
    timeback_pct = a.get("timeback_license_pct", 20) / 100
    upfront_ip = a.get("upfront_ip_fee", 25) * 1_000_000
    capex_per_school = a.get("capex_per_school", 5_000_000)
    avg_per_school = a.get("avg_students_per_school", 800)
    exit_multiple = a.get("exit_ebitda_multiple", 15)
    a.get("discount_rate", 12) / 100
    tax_rate = a.get("tax_rate", 20) / 100
    guide_fee = a.get("guide_school_fee", 15_000)

    # Student counts by year
    student_counts = [
        int(a.get("students_year1", 3000)),
        int(a.get("students_year2", 10000)),
        int(a.get("students_year3", 40000)),
        int(a.get("students_year4", 80000)),
        int(a.get("students_year5", 150000)),
    ]

    # Blended tuition (60% mid, 40% premium)
    blended_tuition = mid_tuition * 0.6 + premium_tuition * 0.4

    # --- P&L Projections ---
    projections: list[YearProjection] = []
    cumulative_cash = 0.0
    prev_schools = 0
    total_mgmt_rev = 0.0
    total_timeback_rev = 0.0

    for yr_idx in range(5):
        year = yr_idx + 1
        students = student_counts[yr_idx]
        schools = max(1, math.ceil(students / avg_per_school))
        new_schools = max(0, schools - prev_schools)

        revenue = students * blended_tuition
        cogs = revenue * cogs_pct
        gross_margin = revenue - cogs
        opex = revenue * opex_pct
        ebitda = gross_margin - opex
        net_income = ebitda * (1 - tax_rate)

        capex = new_schools * capex_per_school
        fcf = net_income - capex
        cumulative_cash += fcf

        mgmt_rev = revenue * mgmt_fee_pct
        timeback_rev = students * per_student_budget * timeback_pct
        total_mgmt_rev += mgmt_rev
        total_timeback_rev += timeback_rev

        projections.append(YearProjection(
            year=year, students=students, schools=schools,
            revenue=round(revenue), cogs=round(cogs),
            gross_margin=round(gross_margin), opex=round(opex),
            ebitda=round(ebitda), net_income=round(net_income),
            free_cash_flow=round(fcf), cumulative_cash=round(cumulative_cash),
        ))
        prev_schools = schools

    # --- Unit Economics ---
    unit_econ = [
        UnitEconomics(
            school_type="Premium",
            per_student_revenue=premium_tuition,
            per_student_cost=round(premium_tuition * cogs_pct),
            contribution_margin=round(premium_tuition * (1 - cogs_pct)),
            margin_pct=round((1 - cogs_pct) * 100, 1),
        ),
        UnitEconomics(
            school_type="Mid-Market",
            per_student_revenue=mid_tuition,
            per_student_cost=round(mid_tuition * cogs_pct),
            contribution_margin=round(mid_tuition * (1 - cogs_pct)),
            margin_pct=round((1 - cogs_pct) * 100, 1),
        ),
    ]

    # --- Capital Deployment ---
    cap_deploy: list[CapitalDeployment] = []
    for yr_idx in range(5):
        year = yr_idx + 1
        students = student_counts[yr_idx]
        schools = max(1, math.ceil(students / avg_per_school))
        new_schools_yr = max(0, schools - (
            0 if yr_idx == 0
            else max(1, math.ceil(student_counts[yr_idx - 1] / avg_per_school))
        ))

        ip_dev = upfront_ip if year == 1 else 0
        mgmt = projections[yr_idx].revenue * mgmt_fee_pct
        launch = new_schools_yr * capex_per_school
        teachers_needed = max(0, students // 20)
        guide_total = (
            teachers_needed * guide_fee if year <= 2
            else teachers_needed * guide_fee * 0.3
        )

        cap_deploy.append(CapitalDeployment(
            year=year, ip_development=round(ip_dev),
            management_fees=round(mgmt), launch_capital=round(launch),
            real_estate=round(launch * 0.6),
            total=round(ip_dev + mgmt + launch + guide_total),
        ))

    # --- Returns Analysis ---
    y5_ebitda = projections[4].ebitda if projections else 0
    ev_exit = y5_ebitda * exit_multiple
    total_invested = sum(cd.total for cd in cap_deploy)
    moic = ev_exit / total_invested if total_invested > 0 else 0

    cash_flows = [-cap_deploy[i].total for i in range(5)]
    cash_flows[4] += ev_exit
    irr = _approx_irr(cash_flows)

    payback = None
    for p in projections:
        if p.cumulative_cash > 0:
            payback = float(p.year)
            break

    returns = ReturnsAnalysis(
        irr=round(irr * 100, 1) if irr else None,
        moic=round(moic, 2),
        enterprise_value_at_exit=round(ev_exit),
        payback_period_years=payback,
        ebitda_multiple=exit_multiple,
    )

    # --- Sensitivity ---
    sensitivity = [
        SensitivityScenario(
            variable="Year 5 Students",
            base_case=student_counts[4],
            downside=student_counts[4] * 0.7,
            upside=student_counts[4] * 1.3,
        ),
        SensitivityScenario(
            variable="Blended Tuition",
            base_case=blended_tuition,
            downside=blended_tuition * 0.85,
            upside=blended_tuition * 1.15,
        ),
        SensitivityScenario(
            variable="COGS %",
            base_case=cogs_pct * 100,
            downside=(cogs_pct + 0.05) * 100,
            upside=(cogs_pct - 0.05) * 100,
        ),
        SensitivityScenario(
            variable="Exit Multiple",
            base_case=exit_multiple,
            downside=exit_multiple * 0.7,
            upside=exit_multiple * 1.3,
        ),
    ]

    return FinancialModel(
        pnl_projection=projections,
        unit_economics=unit_econ,
        capital_deployment=cap_deploy,
        returns_analysis=returns,
        sensitivity=sensitivity,
        management_fee_pct=mgmt_fee_pct,
        timeback_license_pct=timeback_pct,
        upfront_ip_fee=upfront_ip,
        upfront_alphacore_license=a.get("upfront_alphacore_license", 250) * 1_000_000,
        upfront_app_content_rd=a.get("upfront_app_content_rd", 250) * 1_000_000,
        upfront_lifeskills_rd=a.get("upfront_lifeskills_rd", 250) * 1_000_000,
        upfront_mgmt_fee=a.get("upfront_mgmt_fee", 0) * 1_000_000,
        upfront_timeback_fee=a.get("upfront_timeback_fee", 0) * 1_000_000,
        total_management_fee_revenue=round(total_mgmt_rev),
        total_timeback_license_revenue=round(total_timeback_rev),
    )


def recalculate_model(
    current_assumptions: FinancialAssumptions,
    adjustments: dict[str, float],
    target: str,
    strategy: Strategy,
) -> tuple[FinancialAssumptions, FinancialModel]:
    """Apply user adjustments and recalculate the model."""
    for item in current_assumptions.assumptions:
        if item.key in adjustments:
            new_val = adjustments[item.key]
            if item.locked:
                logger.warning("Attempted to change locked assumption %s", item.key)
                continue
            item.value = max(item.min_val, min(item.max_val, new_val))
    model = build_model(current_assumptions, target, strategy)
    return current_assumptions, model


# ---------------------------------------------------------------------------
# Excel export — professionally formatted multi-sheet workbook
# Writes actual computed values (no formula-only cells) for full compatibility
# ---------------------------------------------------------------------------

def export_model_xlsx(
    target: str,
    model: FinancialModel,
    assumptions: FinancialAssumptions,
    country_profile: CountryProfile | None = None,
) -> str:
    """Export the financial model to a professionally formatted Excel workbook.

    Writes actual computed values (not formula-only) so the workbook displays
    correctly in any spreadsheet application. Produces a multi-sheet workbook
    with IB-standard formatting.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output_dir = os.path.join(OUTPUT_DIR, target.lower().replace(" ", "_"))
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f"{target.replace(' ', '_')}_financial_model.xlsx")

    wb = Workbook()

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    label_font = Font(bold=True)
    label_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    input_font = Font(bold=True, color="0000FF")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ebitda_font = Font(bold=True, size=11)
    ebitda_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="000000"))

    def header_row(ws, row, values):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def label_row(ws, row, col, value, bold=True):
        cell = ws.cell(row, col, value)
        if bold:
            cell.font = label_font
            cell.fill = label_fill

    # ================================================================
    # Sheet 1: Assumptions
    # ================================================================
    ws = wb.active
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    ws.cell(1, 1, f"Financial Model — {target}").font = Font(bold=True, size=14, color="002060")
    ws.cell(2, 1, "Key Assumptions (Blue = adjustable input)").font = Font(italic=True, color="666666")
    header_row(ws, 4, ["Assumption", "Value", "Min", "Max"])

    row = 5
    current_cat = ""
    for a in assumptions.assumptions:
        if a.category != current_cat:
            current_cat = a.category
            row += 1
            cell = ws.cell(row, 1, current_cat.upper())
            cell.font = subheader_font
            cell.fill = subheader_fill
            for c2 in range(2, 5):
                ws.cell(row, c2).fill = subheader_fill
            row += 1

        ws.cell(row, 1, a.label).font = label_font
        ws.cell(row, 1).fill = label_fill

        # Write actual numeric values
        display_val = a.value / 100 if a.unit == "%" else a.value
        min_val = a.min_val / 100 if a.unit == "%" else a.min_val
        max_val = a.max_val / 100 if a.unit == "%" else a.max_val

        for c, val in [(2, display_val), (3, min_val), (4, max_val)]:
            cell = ws.cell(row, c, val)
            if a.unit == "%":
                cell.number_format = "0.0%"
            elif a.unit in ("$", "$M"):
                cell.number_format = "$#,##0" if a.unit == "$" else '$#,##0"M"'
            elif a.unit == "x":
                cell.number_format = "0.0x"
            else:
                cell.number_format = "#,##0"

        # IB color coding for inputs
        if not a.locked:
            ws.cell(row, 2).font = input_font
            ws.cell(row, 2).fill = input_fill
        row += 1

    # ================================================================
    # Sheet 2: P&L Projection (5-year)
    # ================================================================
    ws2 = wb.create_sheet("P&L Projection")
    ws2.column_dimensions["A"].width = 25
    for i in range(2, 7):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    ws2.cell(1, 1, f"5-Year P&L Projection — {target}").font = Font(bold=True, size=14, color="002060")
    header_row(ws2, 3, ["Metric"] + [f"Year {p.year}" for p in model.pnl_projection])

    pnl_metrics = [
        ("Students", "students", "#,##0", False),
        ("Schools", "schools", "#,##0", False),
        ("", None, None, None),  # spacer
        ("Revenue", "revenue", "$#,##0", False),
        ("COGS", "cogs", "($#,##0)", False),
        ("Gross Margin", "gross_margin", "$#,##0", True),
        ("", None, None, None),  # spacer
        ("Operating Expenses", "opex", "($#,##0)", False),
        ("EBITDA", "ebitda", "$#,##0", "ebitda"),
        ("Net Income", "net_income", "$#,##0", False),
        ("", None, None, None),  # spacer
        ("Free Cash Flow", "free_cash_flow", "$#,##0", False),
        ("Cumulative Cash", "cumulative_cash", "$#,##0", False),
    ]

    r = 4
    for name, key, fmt, is_total in pnl_metrics:
        if key is None:
            r += 1
            continue
        label_row(ws2, r, 1, name)
        for c, proj in enumerate(model.pnl_projection, start=2):
            cell = ws2.cell(r, c, getattr(proj, key))
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            if is_total == "ebitda":
                cell.font = ebitda_font
                cell.fill = ebitda_fill
            elif is_total:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
        r += 1

    # Alpha Revenue Summary
    r += 1
    ws2.cell(r, 1, "ALPHA REVENUE SUMMARY").font = subheader_font
    ws2.cell(r, 1).fill = subheader_fill
    r += 1
    alpha_items = [
        ("Management Fee Revenue (5yr Total)",
         model.total_management_fee_revenue,
         "$#,##0"),
        ("Timeback License Revenue (5yr Total)",
         model.total_timeback_license_revenue,
         "$#,##0"),
        ("Upfront IP Fee",
         model.upfront_ip_fee,
         "$#,##0"),
        ("Total Alpha Revenue (5yr)",
         model.total_management_fee_revenue +
         model.total_timeback_license_revenue +
         model.upfront_ip_fee,
         "$#,##0"),
    ]
    for name, val, fmt in alpha_items:
        label_row(ws2, r, 1, name)
        cell = ws2.cell(r, 2, val)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        if "Total" in name:
            cell.font = ebitda_font
            cell.fill = ebitda_fill
        r += 1

    # ================================================================
    # Sheet 3: Unit Economics
    # ================================================================
    ws3 = wb.create_sheet("Unit Economics")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 15

    ws3.cell(1, 1, "Per-Student Unit Economics").font = Font(bold=True, size=14, color="002060")
    header_row(ws3, 3, ["School Type", "Revenue/Student", "Cost/Student", "Margin/Student", "Margin %"])

    for r, ue in enumerate(model.unit_economics, start=4):
        ws3.cell(r, 1, ue.school_type).font = label_font
        ws3.cell(r, 2, ue.per_student_revenue).number_format = "$#,##0"
        ws3.cell(r, 3, ue.per_student_cost).number_format = "$#,##0"
        cell4 = ws3.cell(r, 4, ue.contribution_margin)
        cell4.number_format = "$#,##0"
        cell4.font = total_font
        cell4.fill = ebitda_fill
        ws3.cell(r, 5, ue.margin_pct / 100).number_format = "0.0%"

    # ================================================================
    # Sheet 4: Capital Deployment
    # ================================================================
    if model.capital_deployment:
        ws_cap = wb.create_sheet("Capital Deployment")
        ws_cap.column_dimensions["A"].width = 30
        for i in range(2, 7):
            ws_cap.column_dimensions[get_column_letter(i)].width = 18

        ws_cap.cell(1, 1, "Capital Deployment Schedule").font = Font(bold=True, size=14, color="002060")
        header_row(ws_cap, 3, ["Item"] + [f"Year {cd.year}" for cd in model.capital_deployment])

        cap_items = [
            ("IP Development", "ip_development"),
            ("Management Fees", "management_fees"),
            ("Launch Capital", "launch_capital"),
            ("Real Estate", "real_estate"),
            ("Total Capital", "total"),
        ]
        r = 4
        for name, key in cap_items:
            label_row(ws_cap, r, 1, name)
            for c, cd in enumerate(model.capital_deployment, start=2):
                cell = ws_cap.cell(r, c, getattr(cd, key))
                cell.number_format = "$#,##0"
                cell.alignment = Alignment(horizontal="right")
                if key == "total":
                    cell.font = total_font
                    cell.fill = total_fill
            r += 1

    # ================================================================
    # Sheet 5: Returns Analysis
    # ================================================================
    ws4 = wb.create_sheet("Returns Analysis")
    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 25

    ws4.cell(1, 1, "Returns & Valuation Analysis").font = Font(bold=True, size=14, color="002060")

    ret = model.returns_analysis
    returns_items = [
        ("KEY METRICS", None, None, True),
        ("IRR (Internal Rate of Return)", ret.irr / 100 if ret.irr else None, "0.0%", False),
        ("MOIC (Multiple on Invested Capital)", ret.moic, "0.0x", False),
        ("Enterprise Value at Exit", ret.enterprise_value_at_exit, "$#,##0", False),
        ("Payback Period", ret.payback_period_years, '0.0" years"', False),
        ("Exit EBITDA Multiple", ret.ebitda_multiple, "0.0x", False),
        ("", None, None, None),
        ("ALPHA FEE SUMMARY", None, None, True),
        ("Management Fee (% of Revenue)", model.management_fee_pct * 100, "0.0%", False),
        ("Timeback License (% of Revenue)", model.timeback_license_pct * 100, "0.0%", False),
        ("Upfront IP/Development Fee", model.upfront_ip_fee, "$#,##0", False),
        ("Total Management Fee Revenue (5yr)", model.total_management_fee_revenue, "$#,##0", False),
        ("Total Timeback License Revenue (5yr)", model.total_timeback_license_revenue, "$#,##0", False),
    ]

    r = 3
    for name, val, fmt, is_header in returns_items:
        if is_header is None:
            r += 1
            continue
        if is_header:
            ws4.cell(r, 1, name).font = subheader_font
            ws4.cell(r, 1).fill = subheader_fill
            ws4.cell(r, 2).fill = subheader_fill
        else:
            label_row(ws4, r, 1, name)
            if val is not None:
                cell = ws4.cell(r, 2, val)
                cell.number_format = fmt
                cell.font = Font(bold=True, size=12)
                cell.fill = ebitda_fill
            else:
                ws4.cell(r, 2, "N/A")
        r += 1

    # ================================================================
    # Sheet 6: Sensitivity Analysis
    # ================================================================
    ws5 = wb.create_sheet("Sensitivity Analysis")
    ws5.column_dimensions["A"].width = 25
    for col_letter in "BCDE":
        ws5.column_dimensions[col_letter].width = 18

    ws5.cell(1, 1, "Sensitivity Analysis").font = Font(bold=True, size=14, color="002060")
    header_row(ws5, 3, ["Variable", "Base Case", "Downside", "Upside"])

    for r, s in enumerate(model.sensitivity, start=4):
        ws5.cell(r, 1, s.variable).font = label_font
        ws5.cell(r, 2, s.base_case).number_format = "#,##0"
        ws5.cell(r, 3, s.downside).number_format = "#,##0"
        ws5.cell(r, 4, s.upside).number_format = "#,##0"

    # ================================================================
    # Save
    # ================================================================
    wb.save(path)
    wb.close()
    logger.info("Exported financial model XLSX: %s (6 sheets, all computed values)", path)
    return path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _approx_irr(cash_flows: list[float], guess: float = 0.15) -> float | None:
    """Newton-Raphson IRR approximation."""
    rate = guess
    for _ in range(200):
        npv = sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))
        d_npv = sum(-t * cf / (1 + rate) ** (t + 1) for t, cf in enumerate(cash_flows))
        if abs(d_npv) < 1e-12:
            break
        new_rate = rate - npv / d_npv
        if abs(new_rate - rate) < 1e-8:
            return new_rate
        rate = new_rate
    return rate if -1 < rate < 10 else None
